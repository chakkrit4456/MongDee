"""Build one .xlsx workbook covering every logged table, for the Dashboard's
"download all data for this booth/event" button — raw rows, one sheet per
table, so the numbers can be pivoted/charted/analyzed offline in Excel
without going back through the API.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from core import database as db

EXPORT_ROW_LIMIT = 1_000_000  # effectively "all rows" for a booth's lifetime

SHEETS = [
    ("Interactions", db.query_interactions),
    ("HealthEvents", db.query_health_events),
    ("ReadinessChecks", db.query_readiness_checks),
    ("Heartbeats", db.query_heartbeats),
    ("ProductHoldEvents", db.query_product_hold_events),
    ("PresenceSessions", db.query_presence_sessions),
]


def build_workbook(db_path, event_id: str | None = None, booth_id: str | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    for sheet_name, query_fn in SHEETS:
        rows = query_fn(db_path, event_id, booth_id, limit=EXPORT_ROW_LIMIT)
        ws = wb.create_sheet(sheet_name)
        if not rows:
            ws.append(["ไม่มีข้อมูล"])
            continue

        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(h) for h in headers])

        for col_cells in ws.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
